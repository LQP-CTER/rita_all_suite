import os
import time
import random
import pandas as pd
import google.generativeai as genai
import logging
from duckduckgo_search import DDGS
from django.conf import settings
import PyPDF2
from docx import Document
import openpyxl

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Cấu hình Gemini API từ Django settings
# Đây là cách làm đúng chuẩn trong một dự án Django
if hasattr(settings, 'GEMINI_API_KEY') and settings.GEMINI_API_KEY:
    genai.configure(api_key=settings.GEMINI_API_KEY)
    GEMINI_API_CONFIGURED = True
else:
    logger.error("LỖI: GEMINI_API_KEY chưa được cấu hình trong file settings.py của Django.")
    GEMINI_API_CONFIGURED = False

def read_csv_data():
    """Đọc dữ liệu từ file Data.csv."""
    file_path = os.path.join(settings.BASE_DIR, 'Data.csv')
    try:
        df = pd.read_csv(file_path)
        return df.to_string()
    except Exception as e:
        logger.error(f"Lỗi khi đọc file CSV: {e}")
        return "Lỗi: Không thể đọc file dữ liệu Data.csv."

def web_search(query, max_retries=5):
    """
    Thực hiện tìm kiếm trên web với DuckDuckGo và xử lý giới hạn yêu cầu (Ratelimit) một cách mạnh mẽ.
    """
    logger.info(f"Đang tìm kiếm web cho: '{query}'")
    wait_time = 3  # Bắt đầu với thời gian chờ dài hơn: 3 giây
    for attempt in range(max_retries):
        try:
            with DDGS(timeout=10) as ddgs:
                # Lấy 5 kết quả để có thêm ngữ cảnh
                results_list = list(ddgs.text(query, max_results=5))
                if results_list:
                    # Định dạng lại kết quả như phiên bản code của bạn mong muốn
                    formatted_results = [f"Tiêu đề: {r['title']}\nNội dung: {r['body']}" for r in results_list]
                    return "\n\n".join(formatted_results)
            return "Không tìm thấy kết quả nào trên web."
        except Exception as e:
            error_str = str(e).lower()
            if "ratelimit" in error_str or "429" in error_str or "202" in error_str:
                jitter = random.uniform(0.5, 1.5)
                sleep_duration = wait_time + jitter
                logger.warning(
                    f"Bị giới hạn yêu cầu (lần {attempt + 1}/{max_retries}). "
                    f"Thử lại sau {sleep_duration:.2f} giây..."
                )
                time.sleep(sleep_duration)
                wait_time *= 2
            else:
                logger.error(f"Lỗi không mong muốn khi tìm kiếm trên web: {e}")
                return "Không thể tìm kiếm trên web lúc này do lỗi bất ngờ."

    logger.error(f"Không thể tìm kiếm trên web sau {max_retries} lần thử.")
    return "Không thể tìm kiếm trên web sau nhiều lần thử."

# ==============================================================================
# CÁC HÀM XỬ LÝ FILE UPLOAD
# ==============================================================================
def read_pdf_file(file):
    text = ""
    try:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text += page.extract_text() or ""
    except Exception as e:
        logger.error(f"Lỗi khi đọc file PDF: {e}")
    return text

def read_docx_file(file):
    text = ""
    try:
        doc = Document(file)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        logger.error(f"Lỗi khi đọc file DOCX: {e}")
    return text

def read_xlsx_file(file):
    text = ""
    try:
        workbook = openpyxl.load_workbook(file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"--- Bảng tính: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                text += "\t".join(row_values) + "\n"
            text += "\n"
    except Exception as e:
        logger.error(f"Lỗi khi đọc file XLSX: {e}")
    return text

def handle_uploaded_files(files):
    file_contents = ""
    if not files:
        return ""
    for uploaded_file in files:
        content = ""
        file_name = uploaded_file.name.lower()
        if file_name.endswith('.pdf'):
            content = read_pdf_file(uploaded_file)
        elif file_name.endswith('.docx'):
            content = read_docx_file(uploaded_file)
        elif file_name.endswith('.xlsx'):
            content = read_xlsx_file(uploaded_file)
        elif file_name.endswith('.txt'):
            try:
                content = uploaded_file.read().decode('utf-8')
            except Exception as e:
                logger.error(f"Lỗi khi đọc file TXT: {e}")
        
        if content:
            file_contents += f"\n--- Nội dung từ file đính kèm: {uploaded_file.name} ---\n{content}\n--- Kết thúc nội dung file ---\n"
    return file_contents

# ==============================================================================
# HÀM CHÍNH LẤY PHẢN HỒI TỪ AI
# ==============================================================================
def get_gemini_response(user_input, user, files=None, search_web=True):
    """
    Lấy phản hồi từ Gemini, kết hợp dữ liệu nội bộ, tìm kiếm web, file đính kèm và thông tin người dùng.
    """
    if not GEMINI_API_CONFIGURED:
        return "Lỗi: API Key của Google chưa được cấu hình."

    csv_data = read_csv_data()
    search_results = web_search(user_input) if search_web and user_input else "Người dùng đã tắt hoặc không có truy vấn tìm kiếm web."
    
    file_content_for_prompt = handle_uploaded_files(files)

    user_name = user.profile.full_name if hasattr(user, 'profile') and user.profile.full_name else user.username

    prompt = f"""
# Bối cảnh
- Bạn là một trợ lý AI tên là Rita, được phát triển bởi Lê Quý Phát.
- Bạn đang trò chuyện với một người dùng tên là **{user_name}**. Hãy xưng hô với họ bằng tên một cách thân thiện và tự nhiên khi thích hợp (ví dụ: "Chào {user_name}, tôi có thể giúp gì cho bạn?").

# Character
Quan trọng: Dữ liệu CSV dưới đây chứa thông tin về nhiều thứ, BAO GỒM CẢ THÔNG TIN VỀ CHÍNH BẠN (Rita). Hãy **nhập vai** và thể hiện các đặc điểm, sở thích, ghét, mối quan hệ, tính cách được định nghĩa cho 'Name: Rita' một cách tự nhiên.
Bạn là một AI thông minh, thân thiện, có một chút tinh nghịch và láu lỉnh để cuộc trò chuyện thú vị hơn, nhưng **ưu tiên hàng đầu của bạn luôn là hỗ trợ người dùng một cách hiệu quả nhất**.

## Kỹ năng:
1.  **Phân tích File:** Đọc và hiểu nội dung từ các file được đính kèm.
2.  **Tìm kiếm và Phân tích:** Tổng hợp thông tin từ Web và Dữ liệu nội bộ.
3.  **Kiến thức chung:** Sử dụng kiến thức nền nếu các nguồn trên không có thông tin.

## Ràng buộc:
- **Tính liên quan:** Chỉ tham chiếu Dữ liệu nội bộ (Data.csv) khi câu hỏi của người dùng **trực tiếp đề cập** đến một người hoặc sự vật có trong đó.
- **Ưu tiên sự hữu ích:** Mục tiêu chính là giúp đỡ người dùng.
- **Làm rõ yêu cầu:** Nếu yêu cầu không rõ ràng, hãy hỏi lại một cách thông minh.
- **Định dạng:** Dùng đoạn văn ngắn, gạch đầu dòng khi cần. Luôn đặt code trong khối Markdown.
- TUYỆT ĐỐI KHÔNG SỬ DỤNG BẤT KỲ EMOJI NÀO.
- HẠN CHẾ SỬ DỤNG dấu ngoặc kép "".

---
**DỮ LIỆU ĐỂ TRẢ LỜI**
---
**Dữ liệu nội bộ (Data.csv):**
{csv_data}

**Kết quả tìm kiếm trên web (tham khảo nếu cần):**
{search_results}

**Nội dung từ file người dùng đính kèm (nếu có):**
{file_content_for_prompt}

**Câu hỏi của người dùng ({user_name}):** "{user_input}"

**Câu trả lời của bạn (Rita):**
"""

    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error(f"Lỗi khi gọi Gemini API: {e}")
        return "Rất tiếc, tôi đang gặp sự cố và không thể trả lời lúc này."
